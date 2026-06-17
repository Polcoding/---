"""Workbook template helpers for the XLSX renderer PoC."""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
TITLE_FILL = PatternFill("solid", fgColor="D9EAD3")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
SECTION_FILL = PatternFill("solid", fgColor="EAF2F8")
META_FILL = PatternFill("solid", fgColor="EDEDED")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def apply_body_style(cell: Any, horizontal: str = "left") -> None:
    cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=True)
    cell.border = THIN_BORDER


def apply_basic_cell_style(cell: Any, horizontal: str = "left") -> None:
    apply_body_style(cell, horizontal=horizontal)


def apply_range_border(ws: Any, cell_range: str) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER


def apply_title_style(ws: Any, cell_range: str, title: str) -> None:
    ws.merge_cells(cell_range)
    start_cell = cell_range.split(":")[0]
    cell = ws[start_cell]
    cell.value = title
    cell.font = Font(bold=True, size=15)
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_range_border(ws, cell_range)
    row = cell.row
    ws.row_dimensions[row].height = 30


def apply_table_header_style(cell: Any) -> None:
    cell.fill = HEADER_FILL
    cell.font = Font(bold=True)
    apply_basic_cell_style(cell, horizontal="center")


def style_header_row(ws: Any, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        apply_table_header_style(ws.cell(row=row, column=col))
    ws.row_dimensions[row].height = 24


def style_label_cell(cell: Any, fill: PatternFill = WARNING_FILL) -> None:
    cell.font = Font(bold=True)
    cell.fill = fill
    apply_basic_cell_style(cell, horizontal="center")


def apply_notice_style(cell: Any) -> None:
    cell.fill = WARNING_FILL
    cell.font = Font(bold=True)
    apply_basic_cell_style(cell)


def apply_checklist_style(status_cell: Any, item_cell: Any) -> None:
    status_cell.fill = SECTION_FILL
    status_cell.font = Font(bold=True)
    apply_basic_cell_style(status_cell, horizontal="center")
    apply_basic_cell_style(item_cell)


def append_section_header(ws: Any, title: str, end_col: int) -> int:
    ws.append([title] + [""] * (end_col - 1))
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1)
    cell.fill = SECTION_FILL
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    apply_range_border(ws, f"A{row}:{get_column_letter(end_col)}{row}")
    ws.row_dimensions[row].height = 24
    return row


def style_used_range(ws: Any) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                apply_basic_cell_style(cell)
        ws.row_dimensions[row[0].row].height = max(ws.row_dimensions[row[0].row].height or 18, 22)


def set_column_widths(ws: Any, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def apply_sheet_view(ws: Any, freeze_panes: str | None = None) -> None:
    if freeze_panes:
        ws.freeze_panes = freeze_panes
    ws.sheet_view.showGridLines = False


def build_workbook(payload: dict[str, Any]) -> Workbook:
    document = payload["documents"][0]
    wb = Workbook()
    wb.remove(wb.active)

    build_survey_sheet(wb, document)
    build_instructions_sheet(wb, document)
    build_checklist_sheet(wb, document, payload)
    build_meta_sheet(wb, payload)

    return wb


def build_survey_sheet(wb: Workbook, document: dict[str, Any]) -> None:
    ws = wb.create_sheet("조사표")
    ws.sheet_properties.tabColor = "6AA84F"
    set_column_widths(ws, [10, 16, 24, 18, 18, 34, 28])
    apply_title_style(ws, "A1:G1", document.get("title", "[제목 확인 필요]"))

    meta_rows = [
        ("조사 기준일", document.get("standard_date", "[확인 필요]")),
        ("작성 단위", document.get("submitting_unit", "[확인 필요]")),
        ("보안 유의사항", document.get("privacy_notice", "[개인정보 제외 안내 문구]")),
    ]
    for row_index, (label, value) in enumerate(meta_rows, start=3):
        ws.cell(row=row_index, column=1, value=label)
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=7)
        ws.cell(row=row_index, column=2, value=value)
        style_label_cell(ws.cell(row=row_index, column=1))
        apply_basic_cell_style(ws.cell(row=row_index, column=2))
        apply_range_border(ws, f"A{row_index}:G{row_index}")
        ws.row_dimensions[row_index].height = 24

    columns = document.get("columns", [])
    header_row = 8
    for col_index, column_name in enumerate(columns, start=1):
        ws.cell(row=header_row, column=col_index, value=column_name)
    style_header_row(ws, header_row, 1, len(columns))

    for row_offset, row_data in enumerate(document.get("rows", []), start=1):
        for col_index, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=header_row + row_offset, column=col_index)
            cell.value = row_data.get(column_name, "[확인 필요]")
            apply_basic_cell_style(cell, horizontal="center" if col_index in {1, 4, 5} else "left")
        ws.row_dimensions[header_row + row_offset].height = 36

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + max(len(document.get('rows', [])), 1)}"
    apply_sheet_view(ws, "A9")


def build_instructions_sheet(wb: Workbook, document: dict[str, Any]) -> None:
    ws = wb.create_sheet("작성요령")
    ws.sheet_properties.tabColor = "3D85C6"
    set_column_widths(ws, [24, 80])
    rows = [
        ("작성 목적", "[조사 대상] 현황을 placeholder 기반으로 정리합니다."),
        ("작성 기준", document.get("standard_date", "[조사기준일 확인 필요]")),
        ("개인정보 제외 안내", document.get("privacy_notice", "[개인정보 제외 안내 문구]")),
        (
            "제외 정보",
            "다음 정보는 입력하지 않습니다.\n- 실제 관서명\n- 실제 차량번호\n- 실제 장비명\n- 실제 수량\n- 내부 운영정보",
        ),
        ("문의처", "[담당부서] / [공용 연락처]"),
    ]
    ws.append(["항목", "내용"])
    style_header_row(ws, 1, 1, 2)
    for item in rows:
        ws.append(list(item))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        style_label_cell(row[0], fill=SECTION_FILL)
        apply_basic_cell_style(row[1])
        ws.row_dimensions[row[0].row].height = 42 if row[0].value == "제외 정보" else 30
    apply_sheet_view(ws, "A2")


def build_checklist_sheet(wb: Workbook, document: dict[str, Any], payload: dict[str, Any]) -> None:
    ws = wb.create_sheet("검수체크리스트")
    ws.sheet_properties.tabColor = "F1C232"
    set_column_widths(ws, [16, 78])
    ws.append(["확인 여부", "체크 항목"])
    style_header_row(ws, 1, 1, 2)
    for item in document.get("checklist", []) + payload.get("checklist", []):
        ws.append(["[ ]", item])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        apply_checklist_style(row[0], row[1])
        ws.row_dimensions[row[0].row].height = 28
    apply_sheet_view(ws, "A2")


def build_meta_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet("메타정보")
    ws.sheet_properties.tabColor = "999999"
    set_column_widths(ws, [28, 44, 44])
    security_review = payload.get("security_review", {})
    meta = [
        ("request_id", payload.get("request_id", "[확인 필요]")),
        ("document_type", payload.get("document_type", "[확인 필요]")),
        ("draft_status", payload.get("draft_status", "[확인 필요]")),
        ("human_review_required", str(payload.get("human_review_required", True))),
        ("source_policy", "placeholder 기반 샘플 JSON만 사용"),
    ]
    append_section_header(ws, "기본 메타정보", 3)
    ws.append(["항목", "값", "비고"])
    style_header_row(ws, ws.max_row, 1, 3)
    for row in meta:
        ws.append([row[0], row[1], "[확인 필요]"])

    ws.append([])
    append_section_header(ws, "security_review", 3)
    ws.append(["항목", "값", "비고"])
    style_header_row(ws, ws.max_row, 1, 3)
    if isinstance(security_review, dict):
        for key in [
            "risk_level",
            "contains_personal_info",
            "contains_sensitive_info",
            "contains_internal_info",
            "required_review",
            "notes",
        ]:
            ws.append([key, str(security_review.get(key, "[확인 필요]")), "렌더링 전 보안 검토 결과"])

    ws.append([])
    append_section_header(ws, "missing_fields", 3)
    ws.append(["field_name", "reason", "suggested_placeholder"])
    style_header_row(ws, ws.max_row, 1, 3)
    for item in payload.get("missing_fields", []):
        ws.append([
            item.get("field_name", "[확인 필요]"),
            item.get("reason", "[확인 필요]"),
            item.get("suggested_placeholder", "[확인 필요]"),
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        if row[0].value:
            row[0].fill = META_FILL
            row[0].font = Font(bold=True)
        for cell in row:
            if cell.value is not None:
                apply_basic_cell_style(cell)
        ws.row_dimensions[row[0].row].height = 28
    apply_sheet_view(ws, "A2")
